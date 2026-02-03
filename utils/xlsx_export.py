"""Utilitaire pour exporter les réponses contenant des formules Excel dans un fichier .xlsx fonctionnel"""
import re
from io import BytesIO
from typing import Dict, List, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def detect_excel_content(text: str) -> Dict[str, bool]:
    """
    Détecte si le texte contient des formules Excel ou du code VBA.

    Args:
        text: Texte à analyser

    Returns:
        Dict avec les clés 'has_formulas', 'has_vba', 'has_tables'
    """
    result = {
        'has_formulas': False,
        'has_vba': False,
        'has_tables': False
    }

    # Détection de formules Excel (commence par =)
    formula_pattern = r'=\s*[A-Z]+\s*\('
    if re.search(formula_pattern, text, re.IGNORECASE):
        result['has_formulas'] = True

    # Détection de code VBA
    vba_keywords = ['Sub ', 'Function ', 'End Sub', 'End Function', 'Dim ', 'As ', 'Range(', 'Worksheets(']
    if any(keyword in text for keyword in vba_keywords):
        result['has_vba'] = True

    # Détection de tableaux (markdown ou format similaire)
    if '|' in text and text.count('|') > 3:
        result['has_tables'] = True

    return result


def extract_formulas_with_context(text: str) -> List[Dict]:
    """
    Extrait les formules Excel avec leur contexte (description, cellule cible).

    Args:
        text: Texte contenant des formules

    Returns:
        Liste de dict avec 'formula', 'description', 'cell'
    """
    formulas = []
    lines = text.split('\n')

    for i, line in enumerate(lines):
        # Chercher les formules (commence par =)
        if '=' in line and not line.strip().startswith('#'):
            # Extraire la formule
            formula_match = re.search(r'(=[^=\n]+)', line)
            if formula_match:
                formula = formula_match.group(1).strip()
                # Nettoyer les caractères markdown
                formula = formula.replace('`', '').replace('*', '').strip()

                # Si la formule se termine par des caractères markdown, les enlever
                formula = re.sub(r'[`*]+$', '', formula)

                # Chercher une description (ligne avant ou après)
                description = ""
                if i > 0:
                    prev_line = lines[i-1].strip()
                    if prev_line and not prev_line.startswith('='):
                        description = prev_line.replace('*', '').replace('#', '').strip()

                # Chercher si une cellule est mentionnée (ex: "en A1", "dans B2", "B1:")
                cell_match = re.search(r'\b([A-Z]+\d+)\b', line)
                target_cell = cell_match.group(1) if cell_match else None

                formulas.append({
                    'formula': formula,
                    'description': description,
                    'cell': target_cell,
                    'line': line
                })

    return formulas


def parse_formula_examples(text: str) -> List[Dict]:
    """
    Parse les exemples de données pour les formules.
    Cherche les patterns comme:
    - "En A1: valeur" ou "A1 = valeur"
    - "A1: 100"
    """
    examples = []
    lines = text.split('\n')

    # Pattern pour "A1: valeur" ou "A1 = valeur"
    cell_value_pattern = r'([A-Z]+\d+)\s*[:=]\s*(.+)'

    for line in lines:
        match = re.search(cell_value_pattern, line, re.IGNORECASE)
        if match:
            cell = match.group(1).upper()
            value = match.group(2).strip()

            # Nettoyer la valeur
            value = value.replace('`', '').replace('*', '').strip()

            # Ne pas ajouter si c'est une formule
            if value.startswith('='):
                continue

            # Essayer de convertir en nombre
            try:
                if '.' in value or ',' in value:
                    value = float(value.replace(',', '.'))
                else:
                    value = int(value)
            except ValueError:
                # Garder comme texte
                pass

            examples.append({
                'cell': cell,
                'value': value
            })

    return examples


def extract_vba_code(text: str) -> Optional[str]:
    """
    Extrait le code VBA du texte.

    Args:
        text: Texte contenant du code VBA

    Returns:
        Code VBA extrait ou None
    """
    # Chercher les blocs de code VBA (entre ``` ou identifiés par Sub/Function)
    vba_pattern = r'```(?:vba|vb|vbscript)?\n(.*?)```'
    match = re.search(vba_pattern, text, re.DOTALL | re.IGNORECASE)

    if match:
        return match.group(1).strip()

    # Si pas de bloc de code, chercher les Sub/Function
    sub_pattern = r'(Sub\s+\w+.*?End Sub)'
    func_pattern = r'(Function\s+\w+.*?End Function)'

    vba_code = []
    for pattern in [sub_pattern, func_pattern]:
        matches = re.finditer(pattern, text, re.DOTALL | re.IGNORECASE)
        vba_code.extend([match.group(1) for match in matches])

    return '\n\n'.join(vba_code) if vba_code else None


def extract_table_data(text: str) -> Optional[List[List[str]]]:
    """
    Extrait les données tabulaires du texte (format markdown).
    """
    lines = text.split('\n')
    table_data = []
    in_table = False

    for line in lines:
        if '|' in line:
            # Ignorer les lignes de séparation (---|---|---)
            if re.match(r'^\s*\|[\s\-:|]+\|\s*$', line):
                in_table = True
                continue

            # Extraire les cellules
            cells = [cell.strip() for cell in line.split('|')]
            # Retirer les cellules vides au début et à la fin
            cells = [c for c in cells if c]

            if cells:
                table_data.append(cells)
                in_table = True
        elif in_table and table_data:
            # Fin du tableau
            break

    return table_data if table_data else None


def cell_to_coords(cell_ref: str) -> Tuple[int, int]:
    """
    Convertit une référence de cellule (ex: 'A1') en coordonnées (row, col).
    """
    match = re.match(r'([A-Z]+)(\d+)', cell_ref.upper())
    if not match:
        return (1, 1)

    col_str, row_str = match.groups()

    # Convertir la lettre de colonne en numéro
    col = 0
    for char in col_str:
        col = col * 26 + (ord(char) - ord('A') + 1)

    row = int(row_str)

    return (row, col)


def create_excel_export(question: str, response: str, metadata: Dict) -> BytesIO:
    """
    Crée un fichier Excel avec uniquement les feuilles de données, formules et macros.

    Args:
        question: Question posée
        response: Réponse générée
        metadata: Métadonnées (date, modèle, etc.)

    Returns:
        Buffer contenant le fichier Excel
    """
    wb = Workbook()

    # Détection du contenu
    content_types = detect_excel_content(response)

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    normal_font = Font(size=10)
    formula_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    formula_font = Font(bold=True, color="0000FF", size=10)
    code_font = Font(name='Courier New', size=9)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Supprimer la feuille par défaut
    wb.remove(wb.active)

    # === Feuille 1: Feuille principale avec données et formules ===
    ws_main = wb.create_sheet("Feuille1")

    # Extraire les données et formules
    examples = parse_formula_examples(response)
    formulas_data = extract_formulas_with_context(response)
    table_data = extract_table_data(response) if content_types['has_tables'] else None

    # Déterminer la zone d'utilisation maximale
    max_row_used = 0
    max_col_used = 0

    # 1. Insérer les données d'exemple
    if examples:
        for example in examples:
            row, col = cell_to_coords(example['cell'])
            ws_main.cell(row, col, example['value'])
            ws_main.cell(row, col).font = normal_font
            ws_main.cell(row, col).border = border
            max_row_used = max(max_row_used, row)
            max_col_used = max(max_col_used, col)

    # 2. Insérer les tableaux
    if table_data:
        # Trouver une zone libre (après les données existantes)
        start_row = max_row_used + 2

        for row_idx, row_data in enumerate(table_data):
            for col_idx, cell_value in enumerate(row_data, 1):
                current_row = start_row + row_idx
                cell = ws_main.cell(current_row, col_idx, cell_value)
                cell.border = border

                # Essayer de convertir en nombre
                try:
                    if cell_value and cell_value.replace('.', '').replace(',', '').replace('-', '').isdigit():
                        cell.value = float(cell_value.replace(',', '.'))
                except:
                    pass

                # Style pour l'en-tête
                if row_idx == 0:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.font = normal_font

                max_row_used = max(max_row_used, current_row)
                max_col_used = max(max_col_used, col_idx)

    # 3. Insérer les FORMULES ACTIVES
    if formulas_data:
        for formula_data in formulas_data:
            formula = formula_data['formula']
            target_cell = formula_data['cell']

            # Déterminer où placer la formule
            if target_cell:
                row, col = cell_to_coords(target_cell)
            else:
                # Si pas de cellule spécifiée, la placer après les données
                row = max_row_used + 2
                col = 1
                max_row_used = row

            # INSÉRER LA FORMULE RÉELLE
            try:
                ws_main.cell(row, col).value = formula
                ws_main.cell(row, col).font = formula_font
                ws_main.cell(row, col).fill = formula_fill
                ws_main.cell(row, col).border = border

                max_row_used = max(max_row_used, row)
                max_col_used = max(max_col_used, col)

            except Exception as e:
                # Si erreur, mettre un message
                ws_main.cell(row, col, f"ERREUR: {formula}").font = Font(color="FF0000")

    # Ajuster les largeurs de colonnes
    for col in range(1, max_col_used + 1):
        ws_main.column_dimensions[get_column_letter(col)].width = 15

    # Si pas de données du tout, créer une feuille vide avec un message
    if not examples and not formulas_data and not table_data:
        ws_main.cell(1, 1, "Aucune donnée ou formule détectée").font = Font(italic=True)
        ws_main.cell(2, 1, "Vous pouvez utiliser cette feuille pour vos propres données").font = Font(size=9)

    # === Feuille 2: Code VBA (uniquement si détecté) ===
    if content_types['has_vba']:
        ws_vba = wb.create_sheet("Module VBA")

        vba_code = extract_vba_code(response)
        if vba_code:
            ws_vba.cell(1, 1, "Code VBA - Instructions:").font = Font(bold=True, size=12)
            ws_vba.cell(2, 1, "1. Sélectionnez tout le code ci-dessous").font = Font(size=9)
            ws_vba.cell(3, 1, "2. Copiez (Ctrl+C)").font = Font(size=9)
            ws_vba.cell(4, 1, "3. Dans Excel: Alt+F11 → Insert → Module → Collez").font = Font(size=9)
            ws_vba.cell(5, 1, "").font = Font(size=9)

            # Écrire le code
            lines = vba_code.split('\n')
            for idx, line in enumerate(lines, 7):
                ws_vba.cell(idx, 1, line).font = code_font
                ws_vba.cell(idx, 1).alignment = Alignment(wrap_text=False)

        ws_vba.column_dimensions['A'].width = 100

    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
